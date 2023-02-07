"""
Optional: write to Excel worksheet (.xlsx). 

-> This is a simplified application of write_excel. For a more detailed implementation with examples see https://nhsd-git.digital.nhs.uk/data-services/analytics-service/iuod/automated-excel-publications.

Purpose of the script: contains the Excel automation script. Takes the Excel template within excel_templates folder writing output to each worksheet.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
import logging

logger = logging.getLogger(__name__)


class TagNotFound(Exception):
    """
    Exception raised for excel template not containing tags.

    Attributes:
        tag - the tag contained in the Excel template.
        message - the message printed when exception is raised.
        sheet - the Excel sheet missing the tag.

    Outputs:
        Raises an exception if tag is missing from Excel table.
    """

    def __init__(self, tag, sheet): # add sheet
        self.tag = tag
        self.sheet = sheet
        self.message = f"{tag} not found in sheet: {sheet}. Please amend the Excel template."
        super().__init__(self.message)


def find_starting_cell(wb, sheet, tag):
    """Return the row and cell of the tag.

    Inputs:
        wb: the workbook sheet
        sheet: the sheet's name
        tag: the tag in the Excel table template

    Outputs:
        a tuple containing the tag's cell coordinates
    """

    ws = wb[sheet]

    for row in ws.iter_rows(max_row = 1000):
        for cell in row:
            if cell.value == tag:
                return coordinate_to_tuple(cell.coordinate)
    else:
        raise TagNotFound(tag, sheet)


#### User input required: edit the below tables according to your Excel template ####

def prepare_table_1(csv_path):
    logger.info(f"Preparing data for excel table 1")
    # This list sets out the order in which we want the values to appear
    order = ['ALL_ENGLAND',
            'London',
            'South West',
            'South East',
            'Midlands',
            'East of England',
            'North West',
            'North East and Yorkshire',
            'Special Health Authorities and other statutory bodies',]

    logger.info(f"Reading file from:\n{csv_path}")
    df = pd.read_csv(csv_path)
    df_r = df.round(2)
    df_a = df_r[(df_r['BREAKDOWN_TYPE'].isin(['ALL_ENGLAND', 'REGION']))]
    df_a['CONSTANT'] = 'ROW' # This column is only used to help us pivot the table onto one row
    df_p = df_a.pivot_table('COUNT/PERCENT', 'CONSTANT', 'BREAKDOWN_VALUE')

    df_f = df_p[order]

    return df_f


def prepare_table_2(csv_path):
    logger.info(f"Preparing data for excel table 2")
    order = ['BREAKDOWN_1',
        'BREAKDOWN_2',
        'BREAKDOWN_3',
        'BREAKDOWN_4'
    ]

    logger.info(f"Reading file from:\n{csv_path}")
    df = pd.read_csv(csv_path)
    df_r = df.round(2)
    df_a = df_r[(df_r['BREAKDOWN_TYPE'].isin(['BREAKDONW_1', 'BREAKDOWN_2', 'BREAKDOWN_3', 'BREAKDOWN_4', ]))]
    df_a['CONSTANT'] = 'ROW' # This column is only used to help us pivot the table onto one row
    df_p = df_a.pivot_table('COUNT/PERCENT', 'CONSTANT', 'BREAKDOWN_VALUE')
    
    df_f = df_p[order]

    return df_f


def prepare_table_2_1(df):
    logger.info(f"Preparing data for excel tag table 2_1")

    logger.info(f"Reading df:\n{df}")
    df_f = df[["Breakdown_category_totals"]]

    return df_f


def prepare_table_2_2(df):
    logger.info(f"Preparing data for excel tag table 2_2")

    logger.info(f"Reading df:\n{df}")
    df_f = df[['Breakdown_category_2']]

    return df_f


def prepare_table_2_3(df):
    logger.info(f"Preparing data for excel tag table 2_3")

    logger.info(f"Reading df:\n{df}")
    df = df[['Breakdown_category_3', 
        'Subcategory_1',
        'Subcategory_2',
        'Subcategory_3',
        'Subcategory_4'
        'Subcategory_5',
        'Subcategory_6',
        'Subcategory_7',
        'Subcategory_8', 
        'Subcategory_9',
        'Subcategory_10']]
    df_t = df.transpose()

    return df_t


def prepare_table_2_4(df):
    logger.info(f"Preparing data for excel tag table 2_4")

    logger.info(f"Reading df:\n{df}")
    df = df[['Breakdown_category_4',
        'Subcategory_1',
        'Subcategory_2',
        'Subcategory_3',]]
    df_t = df.transpose()

    return df_t

def prepare_table_2_5(df):
    logger.info(f"Preparing data for excel tag table 2_5")

    logger.info(f"Reading df:\n{df}")
    df = df[['Breakdown_category_5',
        'Subcategory_1',
        'Subcategory_2',
        'Subcategory_3',]]
    df_t = df.transpose()

    return df_t


def prepare_table_2_6(df):
    logger.info(f"Preparing data for excel tag table 2_6")

    logger.info(f"Reading df:\n{df}")
    df = df[['Breakdown_category_5',
        'Subcategory_1',
        'Subcategory_2',
        'Subcategory_3',
        'Subcategory_4']]
    df_t = df.transpose()

    return df_t


def prepare_table_2_7(df):
    logger.info(f"Preparing data for excel tag table 2_7")

    logger.info(f"Reading df:\n{df}")
    df_f = df[["Breakdown_categoyry_6"]]

    return df_f


def write_tables_to_excel(tables, excel_template, excel_output):
    """Write data to an excel template.

    Example:
    -------
    write_table_5_4_to_excel(table_data, excel_file=EXCEL_TEMPLATE)
    """
    logger.info(f"Writing tables to excel")
    logger.info(f"Using excel template:\n {excel_template}")
    xl_writer = pd.ExcelWriter(excel_output, engine='openpyxl')
    wb = load_workbook(excel_template)
    xl_writer.book = wb
    xl_writer.sheets = {ws.title: ws for ws in wb.worksheets}

    for table in tables:
        start_cell = find_starting_cell(wb, table["sheet_name"], table["tag"])
        table["data"].to_excel(
            xl_writer,
            table["sheet_name"],
            index=False,
            header=False,
            startcol=start_cell[1]-1,
            startrow=start_cell[0]-1
        )
    logger.info(f"Saving outputs to:\n {excel_output}")
    xl_writer.save()
